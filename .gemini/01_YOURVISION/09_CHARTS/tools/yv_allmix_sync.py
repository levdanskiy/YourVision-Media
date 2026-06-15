#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
YourVision AllMix Vibe Shift 2026 Standing Sync & Card Generator
Parses the Google Sheet results from the AllMix listener round, calculates new ratings,
applies the reach threshold, updates 08_HUB/data.js, and generates Telegram cards.
"""

import openpyxl
import re
import json
import os
import subprocess

EXCEL_PATH = "/home/levdanskiy/Загрузки/YourVision_Chart_Analysis.xlsx"
DATA_JS_PATH = "/home/levdanskiy/GEMINI_PROJECT/01_YOURVISION/08_HUB/data.js"
REBUILD_SCRIPT = "/home/levdanskiy/GEMINI_PROJECT/01_YOURVISION/08_HUB/tools/rebuild_perfect.py"
OUTPUT_CARDS_PATH = "/home/levdanskiy/GEMINI_PROJECT/01_YOURVISION/09_CHARTS/tools/generated_cards.txt"

# Map artists to European flags where relevant
COUNTRY_MAPPING = {
    "linda lampenius": "fi",
    "pete parkkonen": "fi",
    "sal da vinci": "it",
    "alexandra căpitănescu": "ro",
    "zara larsson": "se",
    "dara": "bg",
    "dave": "gb"
}

def get_flag_url(artist):
    artist_lower = artist.lower()
    for key, code in COUNTRY_MAPPING.items():
        if key in artist_lower:
            return f"https://www.eurovision.com/static/images/flags/flag_{code}.svg"
    return "https://www.eurovision.com/static/images/70-heart-sm.ff9bba532601.webp"

def clean_typography(text):
    if not text:
        return ""
    text = text.replace("—", "-").replace("–", "-")
    # Clean up standard straight quotes in Russian
    text = re.sub(r'"([а-яА-ЯёЁ\s]+)"', r'«\1»', text)
    return text

def parse_votes(votes_str):
    votes = []
    if votes_str:
        parts = [p.strip() for p in votes_str.split(",") if p.strip()]
        for p in parts:
            match = re.search(r"([^(]+)\s*\((\d+)\)", p)
            if match:
                voter = match.group(1).strip()
                rank = int(match.group(2).strip())
                votes.append((voter, rank))
    votes.sort(key=lambda x: x[1])
    return votes

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Лист1']
    
    prog_map = {1:15, 2:12, 3:10, 4:9, 5:8, 6:7, 7:6, 8:5, 9:4, 10:3, 11:2, 12:1}
    tracks = []
    all_voters = set()
    
    for r in range(2, 22):
        place = ws.cell(r, 1).value
        track_raw = ws.cell(r, 2).value
        linear_score = ws.cell(r, 3).value
        votes_str = ws.cell(r, 4).value
        
        if not track_raw:
            continue
            
        votes = parse_votes(votes_str)
        for voter, rank in votes:
            all_voters.add(voter)
            
        prog_score = sum(prog_map.get(rank, 0) for voter, rank in votes)
        
        # Split artist and song
        artist = "Unknown Artist"
        song = "Unknown Song"
        if " - " in track_raw:
            parts = track_raw.split(" - ", 1)
            artist = parts[0].strip()
            song = parts[1].strip().strip("«»\"' ")
        else:
            artist = track_raw.strip()
            
        tracks.append({
            "old_place": int(place) if place is not None else None,
            "track_raw": track_raw,
            "artist": artist,
            "song": song,
            "linear_score": int(linear_score) if linear_score is not None else 0,
            "prog_score": prog_score,
            "votes": votes
        })
        
    total_voters = len(all_voters)
    print(f"Total voters: {total_voters}")
    
    # Sort by progressive score descending, then linear score descending
    tracks.sort(key=lambda x: (-x["prog_score"], -x["linear_score"]))
    
    # Assign new places
    for idx, t in enumerate(tracks, 1):
        t["new_place"] = idx
        
    # Helper functions for correct Russian grammar agreements
    def get_points_noun(score):
        last_digit = score % 10
        last_two = score % 100
        if last_two in [11, 12, 13, 14]:
            return "баллов"
        elif last_digit == 1:
            return "балл"
        elif last_digit in [2, 3, 4]:
            return "балла"
        else:
            return "баллов"

    def get_votes_noun(count):
        last_digit = count % 10
        last_two = count % 100
        if last_two in [11, 12, 13, 14]:
            return "голосов"
        elif last_digit == 1:
            return "голос"
        elif last_digit in [2, 3, 4]:
            return "голоса"
        else:
            return "голосов"

    # Generate cards
    cards = []
    for t in tracks:
        vote_count = len(t["votes"])
        reach_pct = int(round((vote_count / total_voters) * 100)) if total_voters > 0 else 0
        rating = (t["linear_score"] / (vote_count * 12)) * 10 if vote_count > 0 else 0.0
        
        badge = ""
        is_outlier = vote_count < (total_voters * 0.25)
        if is_outlier:
            badge = "OUTLIER ⚠️"
        elif t["new_place"] == 1:
            badge = "🏆 ПОБЕДА"
        elif t["prog_score"] >= 50:
            badge = "HIT 🔥"
        elif t["new_place"] <= 5:
            badge = "TOP-5"
        elif t["new_place"] <= 10:
            badge = "TOP-10"
            
        place_emoji = "•"
        if t["new_place"] == 1:
            place_emoji = "🥇"
        elif t["new_place"] == 2:
            place_emoji = "🥈"
        elif t["new_place"] == 3:
            place_emoji = "🥉"
            
        votes_str_formatted = ", ".join([f"{v[0]} ({v[1]})" for v in t["votes"]])
        place_header = f"{place_emoji} {t['new_place']}." if t["new_place"] <= 3 else f"{t['new_place']}."
        
        pts_noun = get_points_noun(t['prog_score'])
        vts_noun = get_votes_noun(vote_count)
        
        card = f"""{place_header} ALLMIX: VIBE SHIFT 2026

🎤 {clean_typography(t['artist'])} - {clean_typography(t['song'])}

📊 {t['prog_score']} {pts_noun} | 👥 {vote_count} {vts_noun} | {badge}
👑 {t['old_place']} | 📈 {reach_pct}% | ⭐️ {rating:.1f}/10

💬 {votes_str_formatted}

#ALLMIX #VibeShift2026"""
        cards.append(card)
        
    # Save cards to generated_cards.txt
    with open(OUTPUT_CARDS_PATH, "w", encoding="utf-8") as f:
        f.write("\n\n---\n\n".join(cards))
    print(f"Cards saved to {OUTPUT_CARDS_PATH}")
    
    # Now build new_chart_standings for data.js
    new_chart_standings = []
    for t in tracks:
        flag_url = get_flag_url(t["artist"])
        new_chart_standings.append({
            "r": t["new_place"],
            "a": clean_typography(t["artist"]),
            "s": clean_typography(t["song"]),
            "p": f"{t['prog_score']} pts (=)",
            "img": flag_url
        })
        
    # Read and update data.js
    with open(DATA_JS_PATH, "r", encoding="utf-8") as f:
        data_content = f.read()
        
    data_match = re.search(r"var DATA = ({.*});", data_content, re.DOTALL)
    if not data_match:
        print("ERROR: Could not parse var DATA object from data.js")
        return
        
    data_obj = json.loads(data_match.group(1), strict=False)
    data_obj["chart"] = new_chart_standings
    
    new_data_json = json.dumps(data_obj, ensure_ascii=False, indent=4)
    new_data_content = f"var DATA = {new_data_json};\n"
    
    with open(DATA_JS_PATH, "w", encoding="utf-8") as f:
        f.write(new_data_content)
    print("Successfully updated DATA.chart inside 08_HUB/data.js.")
    
    # Run rebuild_perfect.py
    if os.path.exists(REBUILD_SCRIPT):
        print(f"Triggering rebuild script: {REBUILD_SCRIPT}...")
        try:
            result = subprocess.run(["python3", REBUILD_SCRIPT], check=True, capture_output=True, text=True)
            print("Hub Rebuild Output:")
            print(result.stdout)
            print("Standings synced and hub rebuilt successfully.")
        except subprocess.CalledProcessError as e:
            print(f"ERROR running rebuild script: {e}")
            print(e.stderr)

if __name__ == "__main__":
    main()
