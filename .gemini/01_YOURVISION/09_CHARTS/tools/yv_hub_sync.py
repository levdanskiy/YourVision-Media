#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
YourVision Hub Standings Sync Tool
Syncs the calculated chart standings from the Excel results spreadsheet into 08_HUB/data.js
and triggers rebuild_perfect.py to update the live dashboard index.html.
"""

import sys
import os
import re
import json
import subprocess
import openpyxl

EXCEL_PATH = "/home/levdanskiy/Загрузки/YourVision_Chart_Analysis.xlsx"
DATA_JS_PATH = "/home/levdanskiy/GEMINI_PROJECT/01_YOURVISION/08_HUB/data.js"
REBUILD_SCRIPT = "/home/levdanskiy/GEMINI_PROJECT/01_YOURVISION/08_HUB/tools/rebuild_perfect.py"

COUNTRY_CODES = {
    "romania": "ro", "australia": "au", "latvia": "lv", "malta": "mt", "czechia": "cz",
    "san marino": "sm", "moldova": "md", "montenegro": "me", "croatia": "hr", "serbia": "rs",
    "portugal": "pt", "switzerland": "ch", "cyprus": "cy", "germany": "de", "greece": "gr",
    "luxembourg": "lu", "denmark": "dk", "norway": "no", "azerbaijan": "az", "ukraine": "ua",
    "finland": "fi", "france": "fr", "georgia": "ge", "italy": "it", "lithuania": "lt",
    "estonia": "ee", "belgium": "be", "albania": "al", "united kingdom": "gb", "netherlands": "nl"
}

def get_country_code(track_name):
    # Example track name: "🇷🇴 Romania: Alexandra Capitanescu - Fuego..."
    track_lower = track_name.lower()
    for country, code in COUNTRY_CODES.items():
        if country in track_lower:
            return code
    return "70" # Default heart logo if not matched

def sync_to_hub(excel_path):
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        return False
    if not os.path.exists(DATA_JS_PATH):
        print(f"ERROR: data.js not found at {DATA_JS_PATH}")
        return False

    # 1. Parse Excel Results
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Результаты']
    
    new_chart_standings = []
    
    for r in range(5, 100):
        new_place = ws.cell(r, 2).value
        delta = ws.cell(r, 3).value
        track_str = ws.cell(r, 4).value
        new_score = ws.cell(r, 6).value
        
        if not track_str:
            continue
            
        # Parse artist and song
        # Format: 🇷🇴 Romania: Alexandra Capitanescu - Fuego (Eleni Foureira, Кипр 2018)
        artist = "Unknown Artist"
        song = "Unknown Song"
        
        # Split flag and country
        parts = track_str.split(":", 1)
        if len(parts) > 1:
            rest = parts[1].strip()
        else:
            rest = track_str.strip()
            
        # Split artist and song by dash
        dash_parts = rest.split(" - ", 1)
        if len(dash_parts) > 1:
            artist = dash_parts[0].strip()
            song = dash_parts[1].strip()
        else:
            artist = rest
            
        code = get_country_code(track_str)
        flag_url = f"https://www.eurovision.com/static/images/flags/flag_{code}.svg"
        if code == "70":
            flag_url = "https://www.eurovision.com/static/images/70-heart-sm.ff9bba532601.webp"
            
        # Format delta
        delta_str = "(=)"
        if delta and delta != "=":
            # If delta has arrow e.g. "↑2" or "↓1", use it
            delta_str = f"({delta})"
            
        new_chart_standings.append({
            "r": int(new_place),
            "a": artist,
            "s": song,
            "p": f"{new_score} pts {delta_str}",
            "img": flag_url
        })
        
    print(f"Parsed {len(new_chart_standings)} chart entries from Excel.")

    # 2. Read and Update data.js
    with open(DATA_JS_PATH, "r", encoding="utf-8") as f:
        data_content = f.read()
        
    data_match = re.search(r"var DATA = ({.*});", data_content, re.DOTALL)
    if not data_match:
        print("ERROR: Could not parse var DATA object from data.js")
        return False
        
    data_obj = json.loads(data_match.group(1), strict=False)
    
    # Update chart array
    data_obj["chart"] = new_chart_standings
    
    # Write back to data.js
    new_data_json = json.dumps(data_obj, ensure_ascii=False, indent=4)
    new_data_content = f"var DATA = {new_data_json};\n"
    
    with open(DATA_JS_PATH, "w", encoding="utf-8") as f:
        f.write(new_data_content)
        
    print("Successfully updated DATA.chart inside 08_HUB/data.js.")

    # 3. Trigger rebuild_perfect.py
    if os.path.exists(REBUILD_SCRIPT):
        print(f"Triggering rebuild script: {REBUILD_SCRIPT}...")
        try:
            result = subprocess.run(["python3", REBUILD_SCRIPT], check=True, capture_output=True, text=True)
            print("Hub Rebuild Output:")
            print(result.stdout)
            print("✅ Standings synced and hub rebuilt successfully.")
            return True
        except subprocess.CalledProcessError as e:
            print(f"ERROR running rebuild script: {e}")
            print(e.stderr)
            return False
    else:
        print(f"WARNING: Rebuild script {REBUILD_SCRIPT} not found.")
        return True

def main():
    excel_file = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    success = sync_to_hub(excel_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
