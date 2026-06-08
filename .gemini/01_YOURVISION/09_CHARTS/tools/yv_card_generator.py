#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
YourVision Charts Telegram Card Generator
Parses Excel results, calculates stats under Hit-Maker and old linear systems,
applies the reach threshold, and outputs publication cards.
"""

import sys
import os
import re
import json
import openpyxl

EXCEL_PATH = "/home/levdanskiy/Загрузки/YourVision_Chart_Analysis.xlsx"

def clean_typography(text):
    if not text:
        return ""
    # Replace long and medium dashes with short hyphens
    text = text.replace("—", "-").replace("–", "-")
    # Replace straight quotes in Russian text with Russian quotes
    text = re.sub(r'"([а-яА-ЯёЁ\s]+)"', r'«\1»', text)
    return text

def parse_excel_results(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: Excel file not found at {filepath}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Результаты']
    
    tracks = []
    all_voters = set()
    
    # Let's read the data rows starting from row 5
    for r in range(5, 100):
        old_place = ws.cell(r, 1).value
        new_place = ws.cell(r, 2).value
        delta = ws.cell(r, 3).value
        track_str = ws.cell(r, 4).value
        old_score = ws.cell(r, 5).value
        new_score = ws.cell(r, 6).value
        jury_votes_str = ws.cell(r, 8).value
        
        if not track_str:
            continue # End of list
            
        # Parse jury votes
        votes = []
        if jury_votes_str:
            parts = [v.strip() for v in jury_votes_str.split(",") if v.strip()]
            for p in parts:
                vote_match = re.search(r"([^(]+)\s*\((\d+)\)", p)
                if vote_match:
                    voter_name = vote_match.group(1).strip()
                    rank_val = int(vote_match.group(2).strip())
                    votes.append((voter_name, rank_val))
                    all_voters.add(voter_name)
                    
        # Sort votes by rank ascending
        votes.sort(key=lambda x: x[1])
        
        tracks.append({
            "old_place": old_place,
            "new_place": new_place,
            "delta": delta,
            "track_raw": track_str,
            "old_score": old_score,
            "new_score": new_score,
            "votes": votes
        })
        
    return tracks, len(all_voters)

def generate_cards(tracks, total_voters):
    cards = []
    
    for track in tracks:
        track_raw = clean_typography(track["track_raw"])
        new_place = track["new_place"]
        old_place = track["old_place"]
        new_score = track["new_score"]
        old_score = track["old_score"]
        votes = track["votes"]
        
        # Parse country flag, artist, song
        # Example format: 🇷🇴 Romania: Alexandra Capitanescu - Fuego (Eleni Foureira, Кипр 2018)
        flag = ""
        artist_song = track_raw
        
        flag_match = re.match(r"^([^\w\s:]+)\s*([^:]+):\s*(.+)$", track_raw)
        if flag_match:
            flag = flag_match.group(1).strip()
            # country_name = flag_match.group(2).strip()
            artist_song = flag_match.group(3).strip()
            
        vote_count = len(votes)
        reach_pct = int(round((vote_count / total_voters) * 100)) if total_voters > 0 else 0
        
        # Rating formula: (Linear Points / (Votes Received * 12)) * 10
        # If Votes Received is 0, rating is 0
        rating = 0.0
        if vote_count > 0:
            rating = (old_score / (vote_count * 12)) * 10
            
        # Determine badges
        badge = ""
        is_outlier = vote_count < (total_voters * 0.25) # Reach Threshold (25%)
        
        if is_outlier:
            badge = "OUTLIER ⚠️"
        elif new_place == 1:
            badge = "🏆 ПОБЕДА"
        elif new_score >= 50:
            badge = "HIT 🔥"
        elif new_place <= 5:
            badge = "TOP-5"
        elif new_place <= 10:
            badge = "TOP-10"
            
        # Place emoji
        place_emoji = "•"
        if new_place == 1:
            place_emoji = "🥇"
        elif new_place == 2:
            place_emoji = "🥈"
        elif new_place == 3:
            place_emoji = "🥉"
            
        # Format votes string: Jane Marduk (1), GLN (3), ...
        votes_str = ", ".join([f"{v[0]} ({v[1]})" for v in votes])
        
        # Place number formatting
        place_header = f"{place_emoji} {new_place}." if new_place <= 3 else f"{new_place}."
        
        card = f"""{place_header} EUROGROOVE: ALBM CUTS

🎤 {flag} {artist_song}

📊 {new_score} баллов | 👥 {vote_count} голосов | {badge}
👑 {old_place} | 📈 {reach_pct}% | ⭐️ {rating:.1f}/10

💬 {votes_str}

#EUROGROOVE #AlbmCuts"""
        cards.append(card)
        
    return cards

def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    
    print(f"Parsing results from: {filepath}...")
    tracks, total_voters = parse_excel_results(filepath)
    print(f"Successfully loaded {len(tracks)} tracks. Total voters detected: {total_voters}\n")
    
    cards = generate_cards(tracks, total_voters)
    
    # Save output to content folder
    output_dir = "/home/levdanskiy/GEMINI_PROJECT/01_YOURVISION/09_CHARTS/tools"
    output_path = os.path.join(output_dir, "generated_cards.txt")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n---\n\n".join(cards))
        
    print(f"✅ Generated {len(cards)} cards and saved them to {output_path}.\n")
    
    # Print the top card as preview
    print("Preview of the 1st card:")
    print("-" * 50)
    print(cards[0])
    print("-" * 50)

if __name__ == "__main__":
    main()
