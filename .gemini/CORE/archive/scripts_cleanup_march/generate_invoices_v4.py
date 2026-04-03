import openpyxl
import datetime
import os
import subprocess
import shutil

# Paths
source_file = "Apsaimniekošana/2026/1_Sistema_Excel/Invoice_System_2026_calc.xlsx"
output_dir = "Apsaimniekošana/2026/1_Sistema_Excel/Rēķini_PDF"

# Create/Clear output dir
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Helper functions
EXCEL_BASE_DATE = datetime.date(1899, 12, 30)

def excel_date_to_str(serial):
    if serial is None: return ""
    if isinstance(serial, (int, float)):
        dt = EXCEL_BASE_DATE + datetime.timedelta(days=serial)
        return dt.strftime("%d.%m.%Y")
    return str(serial)

def excel_date_to_dt_obj(serial):
    if isinstance(serial, (int, float)):
        return EXCEL_BASE_DATE + datetime.timedelta(days=serial)
    return None

def get_float(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0

# 1. READ DATA
print("Reading data...")
wb_data = openpyxl.load_workbook(source_file, data_only=True)
ws_dati = wb_data['DATI']

# Global Data
inv_date_serial = ws_dati['B4'].value
inv_due_serial = ws_dati['B5'].value
inv_period = ws_dati['B6'].value

inv_date_str = excel_date_to_str(inv_date_serial)
inv_due_str = excel_date_to_str(inv_due_serial)
inv_date_obj = excel_date_to_dt_obj(inv_date_serial)
inv_date_fmt_iso = inv_date_obj.strftime("%Y%m%d") if inv_date_obj else "00000000"

# Rates
rates = {
    'electr': get_float(ws_dati['G3'].value),
    'waste': get_float(ws_dati['G4'].value),
    'water': get_float(ws_dati['G5'].value),
    'sewer': get_float(ws_dati['G6'].value),
    'gas_fix': get_float(ws_dati['G7'].value),
    'gas_var': get_float(ws_dati['G8'].value),
    'rent': get_float(ws_dati['G9'].value),
}

# Apartments
apartments = []
for row_idx in range(13, 31):
    def get_val(c): return ws_dati.cell(row=row_idx, column=c).value
    apt_no = get_val(1)
    if apt_no is None: continue
    
    apartments.append({
        'no': apt_no,
        'name': get_val(2),
        'area': get_float(get_val(3)),
        'electr_cnt': get_float(get_val(4)),
        'waste_cnt': get_float(get_val(5)),
        'water_prev': get_val(7), 'water_curr': get_val(8), 'water_diff': get_float(get_val(9)),
        'sewer_prev': get_val(10), 'sewer_curr': get_val(11), 'sewer_diff': get_float(get_val(12)),
        'gas_prev': get_val(13), 'gas_curr': get_val(14), 'gas_diff': get_float(get_val(15)),
        'debt': get_float(get_val(16)),
        'overpay': get_float(get_val(17)),
    })

wb_data.close()
print(f"Loaded {len(apartments)} apartments.")

# 2. GENERATE PDFS
for apt in apartments:
    inv_number = f"{inv_date_fmt_iso}-{int(apt['no']):02d}"
    print(f"Generating Invoice {inv_number}...")
    
    # Load Template
    wb_templ = openpyxl.load_workbook(source_file)
    
    # REMOVE EXTRA SHEETS
    for s_name in wb_templ.sheetnames:
        if s_name != 'RĒĶINS':
            wb_templ.remove(wb_templ[s_name])
    
    ws_inv = wb_templ['RĒĶINS']
    
    # --- Fill Data (Overwrite all formulas) ---
    ws_inv['B1'] = apt['no']
    ws_inv['B3'] = f"Nr. {inv_number}"
    
    # FIX: Date, Due, Period are in Column F (6)
    ws_inv['F3'] = inv_date_str
    ws_inv['F4'] = inv_due_str
    ws_inv['F5'] = inv_period
    
    ws_inv['B10'] = apt['name']
    ws_inv['B11'] = f"Parka iela 24, dz. {int(apt['no'])}, Eleja, Jelgavas novads, LV-3023"

    # Services
    sum_rent = rates['rent'] * apt['area']
    ws_inv['C15'] = rates['rent']; ws_inv['D15'] = apt['area']; ws_inv['F15'] = sum_rent
    
    sum_waste = rates['waste'] * apt['waste_cnt']
    ws_inv['C16'] = rates['waste']; ws_inv['D16'] = apt['waste_cnt']; ws_inv['F16'] = sum_waste
    
    sum_water = rates['water'] * apt['water_diff']
    ws_inv['C17'] = rates['water']; ws_inv['D17'] = apt['water_diff']; ws_inv['F17'] = sum_water
    
    sum_sewer = rates['sewer'] * apt['sewer_diff']
    ws_inv['C18'] = rates['sewer']; ws_inv['D18'] = apt['sewer_diff']; ws_inv['F18'] = sum_sewer
    
    sum_gas_fix = rates['gas_fix'] * 1
    ws_inv['C19'] = rates['gas_fix']; ws_inv['D19'] = 1; ws_inv['F19'] = sum_gas_fix
    
    sum_gas_var = rates['gas_var'] * apt['gas_diff']
    ws_inv['C20'] = rates['gas_var']; ws_inv['D20'] = apt['gas_diff']; ws_inv['F20'] = sum_gas_var
    
    sum_electr = rates['electr'] * apt['electr_cnt']
    ws_inv['C21'] = rates['electr']; ws_inv['D21'] = apt['electr_cnt']; ws_inv['F21'] = sum_electr
    
    # Totals
    total_month = sum_rent + sum_waste + sum_water + sum_sewer + sum_gas_fix + sum_gas_var + sum_electr
    ws_inv['F23'] = total_month
    ws_inv['F24'] = apt['overpay']
    ws_inv['F25'] = apt['debt']
    ws_inv['F26'] = total_month - apt['overpay'] + apt['debt']
    
    # Readings
    ws_inv['C29'] = f"{apt['water_prev']} > {apt['water_curr']}"
    ws_inv['D29'] = f"{apt['water_diff']} m3"
    
    ws_inv['C30'] = f"{apt['sewer_prev']} > {apt['sewer_curr']}"
    ws_inv['D30'] = f"{apt['sewer_diff']} m3"
    
    ws_inv['C31'] = f"{apt['gas_prev']} > {apt['gas_curr']}"
    ws_inv['D31'] = f"{apt['gas_diff']} m3"
    
    # Save Temp Single Sheet
    temp_xlsx = "temp_invoice.xlsx"
    wb_templ.save(temp_xlsx)
    
    # Convert
    cmd = [
        "libreoffice", "--headless", "--convert-to", "pdf", 
        temp_xlsx, "--outdir", output_dir
    ]
    subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    
    # Rename
    default_pdf = os.path.join(output_dir, "temp_invoice.pdf")
    final_name = f"{inv_number}.pdf"
    final_path = os.path.join(output_dir, final_name)
    
    if os.path.exists(default_pdf):
        if os.path.exists(final_path):
             os.remove(final_path)
        shutil.move(default_pdf, final_path)
    
    wb_templ.close()

if os.path.exists("temp_invoice.xlsx"):
    os.remove("temp_invoice.xlsx")

print("Done. Check Rēķini_PDF folder.")
